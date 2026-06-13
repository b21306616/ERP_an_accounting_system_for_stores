"""API v1 routes for counterparties, pricing, purchases, and settlements."""

from __future__ import annotations

import base64
from datetime import date, datetime, time, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import or_
from sqlalchemy.orm import Session, selectinload

from server_app.api.dependencies import get_db
from server_app.api.routers_v1 import error_detail, require_v1_permission, success_response
from server_app.db.models import (
    CashShift,
    Counterparty,
    CounterpartyCategory,
    Currency,
    DebtLedger,
    ExpenseCategory,
    LoyaltyCard,
    LoyaltySetting,
    LoyaltyTransaction,
    Payment,
    PaymentAllocation,
    PriceList,
    PriceListItem,
    Product,
    ProductGroup,
    ProductUom,
    Promotion,
    PurchaseOrder,
    PurchaseOrderLine,
    PurchaseInvoice,
    PurchaseInvoiceLine,
    Sale,
    Service,
    UnitOfMeasure,
    User,
    Warehouse,
)
from server_app.schemas.business_v1 import (
    CounterpartyCategoryCreate,
    CounterpartyCreate,
    CounterpartyUpdate,
    PaymentCreate,
    LoyaltyAdjustmentCreate,
    LoyaltyCardCreate,
    LoyaltyCardUpdate,
    LoyaltySettingsUpdate,
    PriceListCreate,
    PriceListImportPayload,
    PriceListImportRow,
    PriceListItemCreate,
    PromotionCreate,
    PromotionUpdate,
    PurchaseInvoiceCreate,
    PurchaseOrderCreate,
    PurchaseOrderUpdate,
)
from server_app.services.settlements import (
    current_debt_balance,
    generate_doc_number,
    money,
    now_utc,
    post_debt_entry,
    price,
    qty4,
    update_purchase_invoice_payment_status,
)
from server_app.services.warehouse import WarehouseBusinessError, get_or_create_balance, post_stock_movement


router = APIRouter(prefix="/api/v1", tags=["business"])
require_counterparty_view = require_v1_permission("counterparty.view")
require_counterparty_create = require_v1_permission("counterparty.create")
require_counterparty_edit = require_v1_permission("counterparty.edit")
require_counterparty_category = require_v1_permission("counterparty.category_manage")
require_debt_view = require_v1_permission("counterparty.debt_view")
require_payment_create = require_v1_permission("counterparty.payment_create")
require_payment_cancel = require_v1_permission("counterparty.payment_cancel")
require_pricing_view = require_v1_permission("pricing.view")
require_pricing_create = require_v1_permission("pricing.price_list_create")
require_pricing_edit = require_v1_permission("pricing.price_list_edit")
require_price_list_import = require_v1_permission("pricing.price_list_import")
require_price_list_export = require_v1_permission("pricing.price_list_export")
require_promo_manage = require_v1_permission("pricing.promo_manage")
require_loyalty_manage = require_v1_permission("pricing.loyalty_manage")
require_loyalty_adjust = require_v1_permission("pricing.loyalty_adjust")
require_purchase_view = require_v1_permission("purchase.view")
require_purchase_order_create = require_v1_permission("purchase.order_create")
require_purchase_order_edit = require_v1_permission("purchase.order_edit")
require_purchase_order_cancel = require_v1_permission("purchase.order_cancel")
require_purchase_create = require_v1_permission("purchase.invoice_create")
require_purchase_post = require_v1_permission("purchase.post")
require_purchase_cancel = require_v1_permission("purchase.cancel")
require_purchase_return = require_v1_permission("purchase.return")


def _decimal(value: Decimal | int | str | None, places: str) -> str:
    """Return a decimal value as a normalized string."""

    if value is None:
        value = Decimal("0")
    return str(Decimal(value).quantize(Decimal(places)))


def _doc_datetime(doc_date: date) -> datetime:
    """Convert a date document field into a UTC datetime for ledgers."""

    return datetime.combine(doc_date, time.min, tzinfo=timezone.utc)


def _get_or_404(session: Session, model: type[Any], object_id: int, name: str) -> Any:
    """Load one row by primary key or raise a v1 error."""

    item = session.get(model, object_id)
    if item is None:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", f"{name} not found."))
    return item


def _updates(payload: Any) -> dict[str, Any]:
    """Return explicitly supplied update fields."""

    return payload.model_dump(exclude_unset=True)


def _currency_payload(currency: Currency) -> dict[str, Any]:
    """Return a currency payload."""

    return {
        "id": currency.id,
        "code": currency.code,
        "name": currency.name,
        "symbol": currency.symbol,
        "is_system": currency.is_system,
        "is_active": currency.is_active,
    }


def _category_payload(category: CounterpartyCategory) -> dict[str, Any]:
    """Return a counterparty category payload."""

    return {"id": category.id, "name_ru": category.name_ru, "name_tk": category.name_tk}


def _counterparty_payload(session: Session, row: Counterparty, include_debt: bool = False) -> dict[str, Any]:
    """Return a counterparty payload."""

    data = {
        "id": row.id,
        "code": row.code,
        "name": row.name,
        "category_id": row.category_id,
        "category_name_ru": row.category.name_ru if row.category else None,
        "counterparty_type": row.counterparty_type,
        "role_flags": row.role_flags,
        "phone": row.phone,
        "email": row.email,
        "tax_id": row.tax_id,
        "address": row.address,
        "price_list_id": row.price_list_id,
        "price_list_name_ru": row.price_list.name_ru if row.price_list else None,
        "discount_percent": _decimal(row.discount_percent, "0.01"),
        "credit_limit_tmt": _decimal(row.credit_limit_tmt, "0.01"),
        "note": row.note,
        "is_active": row.is_active,
    }
    if include_debt:
        data["debt"] = {
            "receivable": _decimal(current_debt_balance(session, row.id, "receivable"), "0.01"),
            "payable": _decimal(current_debt_balance(session, row.id, "payable"), "0.01"),
        }
    return data


def _price_list_payload(row: PriceList) -> dict[str, Any]:
    """Return a price-list payload."""

    return {
        "id": row.id,
        "name_ru": row.name_ru,
        "name_tk": row.name_tk,
        "currency_id": row.currency_id,
        "currency_code": row.currency.code if row.currency else None,
        "is_default": row.is_default,
        "is_active": row.is_active,
        "note": row.note,
    }


def _price_item_payload(row: PriceListItem) -> dict[str, Any]:
    """Return a price-list item payload."""

    return {
        "id": row.id,
        "price_list_id": row.price_list_id,
        "product_id": row.product_id,
        "product_sku": row.product.sku if row.product else None,
        "product_name": row.product.name if row.product else None,
        "service_id": row.service_id,
        "service_code": row.service.code if row.service else None,
        "service_name_ru": row.service.name_ru if row.service else None,
        "product_uom_id": row.product_uom_id,
        "uom_id": row.uom_id,
        "uom_code": row.uom.code if row.uom else None,
        "price_tmt": _decimal(row.price_tmt, "0.0001"),
        "valid_from": row.valid_from.isoformat() if row.valid_from else None,
        "valid_to": row.valid_to.isoformat() if row.valid_to else None,
    }


def _order_line_payload(line: PurchaseOrderLine) -> dict[str, Any]:
    """Return a purchase order line payload."""

    return {
        "id": line.id,
        "product_id": line.product_id,
        "product_sku": line.product.sku if line.product else None,
        "product_name": line.product.name if line.product else None,
        "service_id": line.service_id,
        "service_code": line.service.code if line.service else None,
        "service_name_ru": line.service.name_ru if line.service else None,
        "expense_category_id": line.expense_category_id,
        "product_uom_id": line.product_uom_id,
        "uom_id": line.uom_id,
        "uom_code": line.uom.code if line.uom else None,
        "quantity_ordered": _decimal(line.quantity_ordered, "0.0001"),
        "quantity_received": _decimal(line.quantity_received, "0.0001"),
        "price_cur": _decimal(line.price_cur, "0.0001"),
        "price_tmt": _decimal(line.price_tmt, "0.0001"),
        "amount_cur": _decimal(line.amount_cur, "0.01"),
        "amount_tmt": _decimal(line.amount_tmt, "0.01"),
    }


def _order_payload(order: PurchaseOrder) -> dict[str, Any]:
    """Return a purchase order payload."""

    return {
        "id": order.id,
        "doc_number": order.doc_number,
        "doc_date": order.doc_date.isoformat() if order.doc_date else None,
        "counterparty_id": order.counterparty_id,
        "counterparty_name": order.counterparty.name if order.counterparty else None,
        "warehouse_id": order.warehouse_id,
        "warehouse_name": order.warehouse.name if order.warehouse else None,
        "currency_id": order.currency_id,
        "currency_code": order.currency.code if order.currency else None,
        "currency_rate": _decimal(order.currency_rate, "0.000001"),
        "total_amount_cur": _decimal(order.total_amount_cur, "0.01"),
        "total_amount_tmt": _decimal(order.total_amount_tmt, "0.01"),
        "status": order.status,
        "note": order.note,
        "sent_by_user_id": order.sent_by_user_id,
        "sent_at": order.sent_at.isoformat() if order.sent_at else None,
        "cancelled_by_user_id": order.cancelled_by_user_id,
        "cancelled_at": order.cancelled_at.isoformat() if order.cancelled_at else None,
        "lines": [_order_line_payload(line) for line in order.lines],
    }


def _promotion_payload(row: Promotion) -> dict[str, Any]:
    """Return a promotion payload."""

    return {
        "id": row.id,
        "name": row.name,
        "promotion_type": row.promotion_type,
        "target_type": row.target_type,
        "product_id": row.product_id,
        "product_sku": row.product.sku if row.product else None,
        "product_group_id": row.product_group_id,
        "product_group_name_ru": row.product_group.name_ru if row.product_group else None,
        "discount_type": row.discount_type,
        "discount_value": _decimal(row.discount_value, "0.0001"),
        "min_quantity": _decimal(row.min_quantity, "0.0001"),
        "gift_product_id": row.gift_product_id,
        "gift_product_sku": row.gift_product.sku if row.gift_product else None,
        "gift_quantity": _decimal(row.gift_quantity, "0.0001"),
        "valid_from": row.valid_from.isoformat() if row.valid_from else None,
        "valid_to": row.valid_to.isoformat() if row.valid_to else None,
        "is_active": row.is_active,
        "note": row.note,
    }


def _loyalty_setting_payload(row: LoyaltySetting) -> dict[str, Any]:
    """Return loyalty settings payload."""

    return {
        "id": row.id,
        "earn_rate_percent": _decimal(row.earn_rate_percent, "0.01"),
        "redemption_limit_percent": _decimal(row.redemption_limit_percent, "0.01"),
        "is_active": row.is_active,
        "note": row.note,
    }


def _loyalty_card_payload(row: LoyaltyCard) -> dict[str, Any]:
    """Return a loyalty card payload."""

    return {
        "id": row.id,
        "card_number": row.card_number,
        "counterparty_id": row.counterparty_id,
        "counterparty_name": row.counterparty.name if row.counterparty else None,
        "owner_name": row.owner_name,
        "phone": row.phone,
        "balance_tmt": _decimal(row.balance_tmt, "0.01"),
        "is_active": row.is_active,
        "note": row.note,
    }


def _loyalty_transaction_payload(row: LoyaltyTransaction) -> dict[str, Any]:
    """Return a loyalty transaction payload."""

    return {
        "id": row.id,
        "loyalty_card_id": row.loyalty_card_id,
        "card_number": row.loyalty_card.card_number if row.loyalty_card else None,
        "transaction_type": row.transaction_type,
        "doc_type": row.doc_type,
        "doc_id": row.doc_id,
        "amount_tmt": _decimal(row.amount_tmt, "0.01"),
        "balance_after": _decimal(row.balance_after, "0.01"),
        "note": row.note,
        "created_by_user_id": row.created_by_user_id,
        "created_at": row.created_at.isoformat() if row.created_at else None,
    }


def _get_loyalty_settings(session: Session) -> LoyaltySetting:
    """Return the singleton loyalty settings row, creating defaults when needed."""

    row = session.query(LoyaltySetting).order_by(LoyaltySetting.id).first()
    if row is None:
        row = LoyaltySetting(earn_rate_percent=Decimal("0"), redemption_limit_percent=Decimal("100"), is_active=True)
        session.add(row)
        session.flush()
    return row


def _post_loyalty_transaction(
    session: Session,
    card: LoyaltyCard,
    *,
    transaction_type: str,
    amount_tmt: Decimal,
    doc_type: str | None,
    doc_id: int | None,
    note: str | None,
    user_id: int | None,
) -> LoyaltyTransaction:
    """Update a loyalty-card balance and append a movement row."""

    amount = money(amount_tmt)
    new_balance = money(card.balance_tmt + amount)
    if new_balance < Decimal("0.00"):
        raise HTTPException(status_code=400, detail=error_detail("INSUFFICIENT_BONUS", "Loyalty card balance is insufficient."))
    card.balance_tmt = new_balance
    transaction = LoyaltyTransaction(
        loyalty_card_id=card.id,
        transaction_type=transaction_type,
        doc_type=doc_type,
        doc_id=doc_id,
        amount_tmt=amount,
        balance_after=new_balance,
        note=note,
        created_by_user_id=user_id,
    )
    session.add(transaction)
    session.flush()
    return transaction


def _resolve_price_import_row(session: Session, row: PriceListImportRow) -> PriceListItemCreate:
    """Resolve product/service codes in an import row to ids."""

    product_id = row.product_id
    service_id = row.service_id
    if product_id is None and row.product_sku:
        product = session.query(Product).filter(Product.sku == row.product_sku).one_or_none()
        if product is None:
            raise HTTPException(status_code=400, detail=error_detail("UNKNOWN_PRODUCT", f"Product not found: {row.product_sku}"))
        product_id = product.id
    if service_id is None and row.service_code:
        service = session.query(Service).filter(Service.code == row.service_code).one_or_none()
        if service is None:
            raise HTTPException(status_code=400, detail=error_detail("UNKNOWN_SERVICE", f"Service not found: {row.service_code}"))
        service_id = service.id
    return PriceListItemCreate(
        product_id=product_id,
        service_id=service_id,
        product_uom_id=row.product_uom_id,
        uom_id=row.uom_id,
        price_tmt=row.price_tmt,
        valid_from=row.valid_from,
        valid_to=row.valid_to,
    )


def _price_rows_from_xlsx(xlsx_base64: str) -> list[PriceListImportRow]:
    """Parse a base64 XLSX workbook into import rows."""

    try:
        workbook = load_workbook(BytesIO(base64.b64decode(xlsx_base64)), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=error_detail("INVALID_XLSX", "Could not parse XLSX workbook.")) from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result: list[PriceListImportRow] = []
    for raw in rows[1:]:
        values = {headers[index]: raw[index] for index in range(min(len(headers), len(raw))) if headers[index]}
        if not any(value is not None and value != "" for value in values.values()):
            continue
        result.append(
            PriceListImportRow(
                product_id=values.get("product_id"),
                product_sku=values.get("product_sku"),
                service_id=values.get("service_id"),
                service_code=values.get("service_code"),
                product_uom_id=values.get("product_uom_id"),
                uom_id=values.get("uom_id"),
                price_tmt=values.get("price_tmt"),
                valid_from=values.get("valid_from"),
                valid_to=values.get("valid_to"),
            )
        )
    return result


def _price_list_xlsx_base64(rows: list[dict[str, Any]]) -> str:
    """Return a base64 XLSX workbook for price-list rows."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "prices"
    headers = [
        "product_id",
        "product_sku",
        "service_id",
        "service_code",
        "product_uom_id",
        "uom_id",
        "price_tmt",
        "valid_from",
        "valid_to",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    stream = BytesIO()
    workbook.save(stream)
    return base64.b64encode(stream.getvalue()).decode("ascii")


def _invoice_line_payload(line: PurchaseInvoiceLine) -> dict[str, Any]:
    """Return a purchase invoice line payload."""

    return {
        "id": line.id,
        "purchase_order_line_id": line.purchase_order_line_id,
        "product_id": line.product_id,
        "product_sku": line.product.sku if line.product else None,
        "product_name": line.product.name if line.product else None,
        "service_id": line.service_id,
        "service_code": line.service.code if line.service else None,
        "service_name_ru": line.service.name_ru if line.service else None,
        "expense_category_id": line.expense_category_id,
        "product_uom_id": line.product_uom_id,
        "uom_id": line.uom_id,
        "uom_code": line.uom.code if line.uom else None,
        "quantity": _decimal(line.quantity, "0.0001"),
        "price_cur": _decimal(line.price_cur, "0.0001"),
        "price_tmt": _decimal(line.price_tmt, "0.0001"),
        "amount_cur": _decimal(line.amount_cur, "0.01"),
        "amount_tmt": _decimal(line.amount_tmt, "0.01"),
        "avg_cost_before": _decimal(line.avg_cost_before, "0.0001") if line.avg_cost_before is not None else None,
        "avg_cost_after": _decimal(line.avg_cost_after, "0.0001") if line.avg_cost_after is not None else None,
    }


def _invoice_payload(invoice: PurchaseInvoice) -> dict[str, Any]:
    """Return a purchase invoice payload."""

    return {
        "id": invoice.id,
        "doc_number": invoice.doc_number,
        "doc_date": invoice.doc_date.isoformat() if invoice.doc_date else None,
        "counterparty_id": invoice.counterparty_id,
        "counterparty_name": invoice.counterparty.name if invoice.counterparty else None,
        "purchase_order_id": invoice.purchase_order_id,
        "warehouse_id": invoice.warehouse_id,
        "warehouse_name": invoice.warehouse.name if invoice.warehouse else None,
        "currency_id": invoice.currency_id,
        "currency_code": invoice.currency.code if invoice.currency else None,
        "currency_rate": _decimal(invoice.currency_rate, "0.000001"),
        "total_amount_cur": _decimal(invoice.total_amount_cur, "0.01"),
        "total_amount_tmt": _decimal(invoice.total_amount_tmt, "0.01"),
        "payment_status": invoice.payment_status,
        "expiry_note": invoice.expiry_note,
        "is_return": invoice.is_return,
        "return_invoice_id": invoice.return_invoice_id,
        "status": invoice.status,
        "note": invoice.note,
        "posted_by_user_id": invoice.posted_by_user_id,
        "posted_at": invoice.posted_at.isoformat() if invoice.posted_at else None,
        "lines": [_invoice_line_payload(line) for line in invoice.lines],
    }


def _debt_payload(row: DebtLedger) -> dict[str, Any]:
    """Return a debt ledger payload."""

    return {
        "id": row.id,
        "counterparty_id": row.counterparty_id,
        "counterparty_name": row.counterparty.name if row.counterparty else None,
        "debt_type": row.debt_type,
        "doc_type": row.doc_type,
        "doc_id": row.doc_id,
        "doc_number": row.doc_number,
        "doc_date": row.doc_date.isoformat() if row.doc_date else None,
        "amount_tmt": _decimal(row.amount_tmt, "0.01"),
        "balance_after": _decimal(row.balance_after, "0.01"),
        "currency_id": row.currency_id,
        "amount_cur": _decimal(row.amount_cur, "0.01") if row.amount_cur is not None else None,
        "note": row.note,
    }


def _payment_payload(payment: Payment) -> dict[str, Any]:
    """Return a payment payload."""

    return {
        "id": payment.id,
        "doc_number": payment.doc_number,
        "doc_date": payment.doc_date.isoformat() if payment.doc_date else None,
        "counterparty_id": payment.counterparty_id,
        "counterparty_name": payment.counterparty.name if payment.counterparty else None,
        "direction": payment.direction,
        "payment_method": payment.payment_method,
        "amount_tmt": _decimal(payment.amount_tmt, "0.01"),
        "currency_id": payment.currency_id,
        "amount_cur": _decimal(payment.amount_cur, "0.01") if payment.amount_cur is not None else None,
        "currency_rate": _decimal(payment.currency_rate, "0.000001") if payment.currency_rate is not None else None,
        "cash_shift_id": payment.cash_shift_id,
        "status": payment.status,
        "note": payment.note,
        "allocations": [
            {
                "id": allocation.id,
                "doc_type": allocation.doc_type,
                "doc_id": allocation.doc_id,
                "allocated_amount": _decimal(allocation.allocated_amount, "0.01"),
            }
            for allocation in payment.allocations
        ],
    }


def _order_query(session: Session):
    """Return a purchase order query with response relationships loaded."""

    return session.query(PurchaseOrder).options(
        selectinload(PurchaseOrder.counterparty),
        selectinload(PurchaseOrder.warehouse),
        selectinload(PurchaseOrder.currency),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.product),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.service),
        selectinload(PurchaseOrder.lines).selectinload(PurchaseOrderLine.uom),
    )


def _refresh_order(session: Session, order_id: int) -> PurchaseOrder:
    """Reload a purchase order with response relationships."""

    order = _order_query(session).filter(PurchaseOrder.id == order_id).one_or_none()
    if order is None:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "Purchase order not found."))
    return order


def _invoice_query(session: Session):
    """Return a purchase invoice query with response relationships loaded."""

    return session.query(PurchaseInvoice).options(
        selectinload(PurchaseInvoice.counterparty),
        selectinload(PurchaseInvoice.purchase_order),
        selectinload(PurchaseInvoice.warehouse),
        selectinload(PurchaseInvoice.currency),
        selectinload(PurchaseInvoice.lines).selectinload(PurchaseInvoiceLine.product),
        selectinload(PurchaseInvoice.lines).selectinload(PurchaseInvoiceLine.service),
        selectinload(PurchaseInvoice.lines).selectinload(PurchaseInvoiceLine.uom),
    )


def _refresh_invoice(session: Session, invoice_id: int) -> PurchaseInvoice:
    """Reload an invoice with response relationships."""

    invoice = _invoice_query(session).filter(PurchaseInvoice.id == invoice_id).one_or_none()
    if invoice is None:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "Purchase invoice not found."))
    return invoice


def _ensure_active_currency(session: Session, currency_id: int) -> Currency:
    """Return an active currency."""

    currency = _get_or_404(session, Currency, currency_id, "Currency")
    if not currency.is_active:
        raise HTTPException(status_code=400, detail=error_detail("INACTIVE_CURRENCY", "Currency is inactive."))
    return currency


def _ensure_active_counterparty(session: Session, counterparty_id: int) -> Counterparty:
    """Return an active counterparty."""

    counterparty = _get_or_404(session, Counterparty, counterparty_id, "Counterparty")
    if not counterparty.is_active:
        raise HTTPException(status_code=400, detail=error_detail("INACTIVE_COUNTERPARTY", "Counterparty is inactive."))
    return counterparty


def _ensure_supplier(counterparty: Counterparty) -> None:
    """Require supplier role for purchase documents."""

    if counterparty.role_flags not in {1, 3} and counterparty.counterparty_type not in {"supplier", "both"}:
        raise HTTPException(status_code=400, detail=error_detail("NOT_SUPPLIER", "Counterparty is not marked as a supplier."))


def _ensure_active_warehouse(session: Session, warehouse_id: int) -> Warehouse:
    """Return an active warehouse."""

    warehouse = _get_or_404(session, Warehouse, warehouse_id, "Warehouse")
    if not warehouse.is_active:
        raise HTTPException(status_code=400, detail=error_detail("INACTIVE_WAREHOUSE", "Warehouse is inactive."))
    return warehouse


def _line_uom_id(session: Session, product: Product | None, product_uom_id: int | None, uom_id: int | None) -> int | None:
    """Resolve the stock UOM used for a product or service line."""

    if product_uom_id is not None:
        product_uom = _get_or_404(session, ProductUom, product_uom_id, "Product UOM")
        return product_uom.uom_id
    if uom_id is not None:
        _get_or_404(session, UnitOfMeasure, uom_id, "Unit of measure")
        return uom_id
    return product.base_uom_id if product is not None else None


def _create_order_line(session: Session, order: PurchaseOrder, line_payload: Any) -> PurchaseOrderLine:
    """Create one purchase order line and validate references."""

    line_qty = qty4(line_payload.quantity)
    line_price_cur = price(line_payload.price_cur)
    line_price_tmt = price(line_price_cur * order.currency_rate)
    amount_cur = money(line_qty * line_price_cur)
    amount_tmt = money(line_qty * line_price_tmt)
    product: Product | None = None

    if line_payload.product_id is not None:
        product = _get_or_404(session, Product, line_payload.product_id, "Product")
        if not product.is_active:
            raise HTTPException(status_code=400, detail=error_detail("INACTIVE_PRODUCT", "Product is inactive."))
    if line_payload.service_id is not None:
        service = _get_or_404(session, Service, line_payload.service_id, "Service")
        if not service.is_active:
            raise HTTPException(status_code=400, detail=error_detail("INACTIVE_SERVICE", "Service is inactive."))
        _get_or_404(session, ExpenseCategory, line_payload.expense_category_id, "Expense category")

    resolved_uom_id = _line_uom_id(session, product, line_payload.product_uom_id, line_payload.uom_id)
    return PurchaseOrderLine(
        product_id=line_payload.product_id,
        service_id=line_payload.service_id,
        expense_category_id=line_payload.expense_category_id,
        product_uom_id=line_payload.product_uom_id,
        uom_id=resolved_uom_id,
        quantity_ordered=line_qty,
        quantity_received=Decimal("0.0000"),
        price_cur=line_price_cur,
        price_tmt=line_price_tmt,
        amount_cur=amount_cur,
        amount_tmt=amount_tmt,
    )


def _recalculate_order_totals(order: PurchaseOrder) -> None:
    """Recalculate purchase order totals from lines."""

    order.total_amount_cur = money(sum((line.amount_cur for line in order.lines), Decimal("0")))
    order.total_amount_tmt = money(sum((line.amount_tmt for line in order.lines), Decimal("0")))


def _update_order_status_from_receipts(order: PurchaseOrder) -> None:
    """Set purchase order status from received quantities."""

    if order.status == "cancelled":
        return
    if not order.lines:
        order.status = "sent" if order.sent_at else "draft"
        return
    received_quantities = [qty4(line.quantity_received) for line in order.lines]
    ordered_quantities = [qty4(line.quantity_ordered) for line in order.lines]
    if all(received >= ordered for received, ordered in zip(received_quantities, ordered_quantities)):
        order.status = "received"
    elif any(received > Decimal("0.0000") for received in received_quantities):
        order.status = "partial"
    else:
        order.status = "sent" if order.sent_at else "draft"


def _apply_invoice_to_order(session: Session, invoice: PurchaseInvoice, direction: int) -> None:
    """Adjust linked purchase-order received quantities for invoice posting/cancel."""

    if invoice.purchase_order_id is None:
        return
    order = session.get(PurchaseOrder, invoice.purchase_order_id)
    if order is None:
        return
    for line in invoice.lines:
        if line.purchase_order_line_id is None:
            continue
        order_line = session.get(PurchaseOrderLine, line.purchase_order_line_id)
        if order_line is None:
            continue
        sign = Decimal(direction)
        if invoice.is_return:
            sign = -sign
        new_quantity = qty4(order_line.quantity_received + (qty4(line.quantity) * sign))
        if new_quantity < Decimal("0.0000"):
            new_quantity = Decimal("0.0000")
        order_line.quantity_received = new_quantity
    session.flush()
    _update_order_status_from_receipts(order)


def _ensure_order_line_matches_invoice(invoice: PurchaseInvoice, order_line: PurchaseOrderLine, line_payload: Any) -> None:
    """Validate that an invoice line belongs to and matches its order line."""

    if invoice.purchase_order_id is None or order_line.purchase_order_id != invoice.purchase_order_id:
        raise HTTPException(status_code=400, detail=error_detail("ORDER_LINE_MISMATCH", "Purchase order line does not belong to the invoice order."))
    if order_line.product_id != line_payload.product_id or order_line.service_id != line_payload.service_id:
        raise HTTPException(status_code=400, detail=error_detail("ORDER_LINE_TARGET_MISMATCH", "Invoice line target differs from the purchase order line."))


def _create_invoice_line(session: Session, invoice: PurchaseInvoice, line_payload: Any) -> PurchaseInvoiceLine:
    """Create one invoice line and validate references."""

    line_qty = qty4(line_payload.quantity)
    line_price_cur = price(line_payload.price_cur)
    line_price_tmt = price(line_price_cur * invoice.currency_rate)
    amount_cur = money(line_qty * line_price_cur)
    amount_tmt = money(line_qty * line_price_tmt)
    product: Product | None = None

    if line_payload.purchase_order_line_id is not None:
        order_line = _get_or_404(session, PurchaseOrderLine, line_payload.purchase_order_line_id, "Purchase order line")
        _ensure_order_line_matches_invoice(invoice, order_line, line_payload)

    if line_payload.product_id is not None:
        product = _get_or_404(session, Product, line_payload.product_id, "Product")
        if not product.is_active:
            raise HTTPException(status_code=400, detail=error_detail("INACTIVE_PRODUCT", "Product is inactive."))
    if line_payload.service_id is not None:
        service = _get_or_404(session, Service, line_payload.service_id, "Service")
        if not service.is_active:
            raise HTTPException(status_code=400, detail=error_detail("INACTIVE_SERVICE", "Service is inactive."))
        _get_or_404(session, ExpenseCategory, line_payload.expense_category_id, "Expense category")

    resolved_uom_id = _line_uom_id(session, product, line_payload.product_uom_id, line_payload.uom_id)
    return PurchaseInvoiceLine(
        purchase_order_line_id=line_payload.purchase_order_line_id,
        product_id=line_payload.product_id,
        service_id=line_payload.service_id,
        expense_category_id=line_payload.expense_category_id,
        product_uom_id=line_payload.product_uom_id,
        uom_id=resolved_uom_id,
        quantity=line_qty,
        price_cur=line_price_cur,
        price_tmt=line_price_tmt,
        amount_cur=amount_cur,
        amount_tmt=amount_tmt,
    )


@router.get("/currencies")
def list_currencies(
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List currencies through API v1."""

    rows = session.query(Currency).order_by(Currency.code).all()
    return success_response([_currency_payload(row) for row in rows])


@router.get("/counterparty-categories")
def list_counterparty_categories(
    _: User = Depends(require_counterparty_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List counterparty categories."""

    rows = session.query(CounterpartyCategory).order_by(CounterpartyCategory.name_ru).all()
    return success_response([_category_payload(row) for row in rows])


@router.post("/counterparty-categories", status_code=status.HTTP_201_CREATED)
def create_counterparty_category(
    payload: CounterpartyCategoryCreate,
    _: User = Depends(require_counterparty_category),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a counterparty category."""

    category = CounterpartyCategory(**payload.model_dump())
    session.add(category)
    session.commit()
    session.refresh(category)
    return success_response(_category_payload(category))


@router.get("/counterparties")
def list_counterparties(
    search: str | None = Query(default=None),
    include_debt: bool = Query(default=False),
    _: User = Depends(require_counterparty_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List counterparties."""

    query = session.query(Counterparty).options(selectinload(Counterparty.category), selectinload(Counterparty.price_list))
    if search:
        pattern = f"%{search}%"
        query = query.filter(or_(Counterparty.code.ilike(pattern), Counterparty.name.ilike(pattern), Counterparty.phone.ilike(pattern)))
    rows = query.order_by(Counterparty.name).limit(500).all()
    return success_response([_counterparty_payload(session, row, include_debt=include_debt) for row in rows])


@router.post("/counterparties", status_code=status.HTTP_201_CREATED)
def create_counterparty(
    payload: CounterpartyCreate,
    _: User = Depends(require_counterparty_create),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a counterparty."""

    if session.query(Counterparty).filter(Counterparty.code == payload.code).one_or_none() is not None:
        raise HTTPException(status_code=409, detail=error_detail("DUPLICATE_CODE", "Counterparty code already exists."))
    if payload.category_id is not None:
        _get_or_404(session, CounterpartyCategory, payload.category_id, "Counterparty category")
    if payload.price_list_id is not None:
        _get_or_404(session, PriceList, payload.price_list_id, "Price list")
    row = Counterparty(**payload.model_dump())
    session.add(row)
    session.commit()
    return success_response(_counterparty_payload(session, row))


@router.patch("/counterparties/{counterparty_id}")
def update_counterparty(
    counterparty_id: int,
    payload: CounterpartyUpdate,
    _: User = Depends(require_counterparty_edit),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Update a counterparty."""

    row = _get_or_404(session, Counterparty, counterparty_id, "Counterparty")
    updates = _updates(payload)
    if updates.get("category_id") is not None:
        _get_or_404(session, CounterpartyCategory, int(updates["category_id"]), "Counterparty category")
    if updates.get("price_list_id") is not None:
        _get_or_404(session, PriceList, int(updates["price_list_id"]), "Price list")
    for key, value in updates.items():
        setattr(row, key, value)
    session.commit()
    return success_response(_counterparty_payload(session, row))


@router.get("/counterparties/{counterparty_id}/debt-summary")
def get_counterparty_debt_summary(
    counterparty_id: int,
    _: User = Depends(require_debt_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Return receivable/payable debt balances for one counterparty."""

    row = _get_or_404(session, Counterparty, counterparty_id, "Counterparty")
    return success_response(_counterparty_payload(session, row, include_debt=True)["debt"])


@router.get("/price-lists")
def list_price_lists(
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List price lists."""

    rows = session.query(PriceList).options(selectinload(PriceList.currency)).order_by(PriceList.name_ru).all()
    return success_response([_price_list_payload(row) for row in rows])


@router.post("/price-lists", status_code=status.HTTP_201_CREATED)
def create_price_list(
    payload: PriceListCreate,
    _: User = Depends(require_pricing_create),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a price list."""

    _ensure_active_currency(session, payload.currency_id)
    if payload.is_default:
        session.query(PriceList).filter(PriceList.is_default.is_(True)).update({PriceList.is_default: False})
    row = PriceList(**payload.model_dump())
    session.add(row)
    session.commit()
    session.refresh(row)
    return success_response(_price_list_payload(row))


@router.post("/price-lists/{price_list_id}/items", status_code=status.HTTP_201_CREATED)
def add_price_list_item(
    price_list_id: int,
    payload: PriceListItemCreate,
    _: User = Depends(require_pricing_edit),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Add one versioned price-list item."""

    price_list = _get_or_404(session, PriceList, price_list_id, "Price list")
    if payload.product_id is not None:
        _get_or_404(session, Product, payload.product_id, "Product")
    if payload.service_id is not None:
        _get_or_404(session, Service, payload.service_id, "Service")
    if payload.product_uom_id is not None:
        _get_or_404(session, ProductUom, payload.product_uom_id, "Product UOM")
    if payload.uom_id is not None:
        _get_or_404(session, UnitOfMeasure, payload.uom_id, "Unit of measure")
    item = PriceListItem(price_list=price_list, **payload.model_dump())
    session.add(item)
    session.commit()
    return success_response(_price_item_payload(item))


@router.get("/price-lists/{price_list_id}/export")
def export_price_list(
    price_list_id: int,
    _: User = Depends(require_price_list_export),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Export a price list as API rows plus a base64 XLSX workbook."""

    price_list = _get_or_404(session, PriceList, price_list_id, "Price list")
    rows = (
        session.query(PriceListItem)
        .options(selectinload(PriceListItem.product), selectinload(PriceListItem.service), selectinload(PriceListItem.uom))
        .filter(PriceListItem.price_list_id == price_list.id)
        .order_by(PriceListItem.product_id, PriceListItem.service_id, PriceListItem.valid_from.desc(), PriceListItem.id.desc())
        .all()
    )
    payload_rows = [_price_item_payload(row) for row in rows]
    return success_response(
        {
            "price_list": _price_list_payload(price_list),
            "rows": payload_rows,
            "format": "xlsx",
            "filename": f"price-list-{price_list.id}.xlsx",
            "xlsx_base64": _price_list_xlsx_base64(payload_rows),
        }
    )


@router.post("/price-lists/{price_list_id}/import")
def import_price_list(
    price_list_id: int,
    payload: PriceListImportPayload,
    _: User = Depends(require_price_list_import),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Import price-list rows from JSON and/or a base64 XLSX workbook."""

    price_list = _get_or_404(session, PriceList, price_list_id, "Price list")
    raw_rows = list(payload.rows)
    if payload.xlsx_base64:
        raw_rows.extend(_price_rows_from_xlsx(payload.xlsx_base64))
    if not raw_rows:
        raise HTTPException(status_code=400, detail=error_detail("EMPTY_IMPORT", "No price rows were supplied."))

    created = 0
    updated = 0
    skipped = 0
    changed_items: list[PriceListItem] = []
    for raw_row in raw_rows:
        item_payload = _resolve_price_import_row(session, raw_row)
        if item_payload.product_id is not None:
            _get_or_404(session, Product, item_payload.product_id, "Product")
        if item_payload.service_id is not None:
            _get_or_404(session, Service, item_payload.service_id, "Service")
        if item_payload.product_uom_id is not None:
            _get_or_404(session, ProductUom, item_payload.product_uom_id, "Product UOM")
        if item_payload.uom_id is not None:
            _get_or_404(session, UnitOfMeasure, item_payload.uom_id, "Unit of measure")
        existing = (
            session.query(PriceListItem)
            .filter(
                PriceListItem.price_list_id == price_list.id,
                PriceListItem.product_id == item_payload.product_id,
                PriceListItem.service_id == item_payload.service_id,
                PriceListItem.product_uom_id == item_payload.product_uom_id,
                PriceListItem.uom_id == item_payload.uom_id,
                PriceListItem.valid_from == item_payload.valid_from,
            )
            .order_by(PriceListItem.id.desc())
            .first()
        )
        if existing is not None and payload.duplicate_mode == "skip":
            skipped += 1
            continue
        if existing is not None and payload.duplicate_mode == "update":
            existing.price_tmt = price(item_payload.price_tmt)
            existing.valid_to = item_payload.valid_to
            changed_items.append(existing)
            updated += 1
            continue
        item = PriceListItem(price_list=price_list, **item_payload.model_dump())
        session.add(item)
        session.flush()
        changed_items.append(item)
        created += 1

    session.commit()
    return success_response(
        {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "rows": [_price_item_payload(row) for row in changed_items],
        }
    )


def _validate_promotion_row(session: Session, row: Promotion) -> None:
    """Validate promotion references and type-specific fields."""

    if row.valid_to is not None and row.valid_to < row.valid_from:
        raise HTTPException(status_code=400, detail=error_detail("INVALID_DATES", "Promotion valid_to cannot be before valid_from."))
    if row.target_type == "product":
        if row.product_id is None:
            raise HTTPException(status_code=400, detail=error_detail("PRODUCT_REQUIRED", "Product promotion requires product_id."))
        _get_or_404(session, Product, row.product_id, "Product")
    if row.target_type == "group":
        if row.product_group_id is None:
            raise HTTPException(status_code=400, detail=error_detail("PRODUCT_GROUP_REQUIRED", "Group promotion requires product_group_id."))
        _get_or_404(session, ProductGroup, row.product_group_id, "Product group")
    if row.promotion_type == "discount":
        if row.discount_type is None:
            raise HTTPException(status_code=400, detail=error_detail("DISCOUNT_TYPE_REQUIRED", "Discount promotion requires discount_type."))
    if row.promotion_type == "gift":
        if row.gift_product_id is None or qty4(row.gift_quantity) <= Decimal("0.0000"):
            raise HTTPException(status_code=400, detail=error_detail("GIFT_REQUIRED", "Gift promotion requires gift_product_id and positive gift_quantity."))
        _get_or_404(session, Product, row.gift_product_id, "Gift product")


@router.get("/promotions")
def list_promotions(
    active_only: bool = Query(default=False),
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List promotion rules."""

    query = session.query(Promotion).options(selectinload(Promotion.product), selectinload(Promotion.product_group), selectinload(Promotion.gift_product))
    if active_only:
        query = query.filter(Promotion.is_active.is_(True))
    rows = query.order_by(Promotion.id.desc()).all()
    return success_response([_promotion_payload(row) for row in rows])


@router.post("/promotions", status_code=status.HTTP_201_CREATED)
def create_promotion(
    payload: PromotionCreate,
    _: User = Depends(require_promo_manage),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a sale promotion rule."""

    row = Promotion(**payload.model_dump())
    _validate_promotion_row(session, row)
    session.add(row)
    session.commit()
    return success_response(_promotion_payload(row))


@router.patch("/promotions/{promotion_id}")
def update_promotion(
    promotion_id: int,
    payload: PromotionUpdate,
    _: User = Depends(require_promo_manage),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Patch a promotion rule."""

    row = _get_or_404(session, Promotion, promotion_id, "Promotion")
    for key, value in _updates(payload).items():
        setattr(row, key, value)
    _validate_promotion_row(session, row)
    session.commit()
    return success_response(_promotion_payload(row))


@router.get("/loyalty-settings")
def get_loyalty_settings_endpoint(
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Return global loyalty settings."""

    return success_response(_loyalty_setting_payload(_get_loyalty_settings(session)))


@router.put("/loyalty-settings")
def update_loyalty_settings(
    payload: LoyaltySettingsUpdate,
    _: User = Depends(require_loyalty_manage),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Update global loyalty settings."""

    row = _get_loyalty_settings(session)
    for key, value in payload.model_dump().items():
        setattr(row, key, value)
    session.commit()
    return success_response(_loyalty_setting_payload(row))


@router.get("/loyalty-cards")
def list_loyalty_cards(
    search: str | None = Query(default=None),
    active_only: bool = Query(default=False),
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List loyalty cards."""

    query = session.query(LoyaltyCard).options(selectinload(LoyaltyCard.counterparty))
    if active_only:
        query = query.filter(LoyaltyCard.is_active.is_(True))
    if search:
        pattern = f"%{search}%"
        query = query.join(Counterparty, LoyaltyCard.counterparty_id == Counterparty.id, isouter=True).filter(
            or_(LoyaltyCard.card_number.ilike(pattern), LoyaltyCard.owner_name.ilike(pattern), LoyaltyCard.phone.ilike(pattern), Counterparty.name.ilike(pattern))
        )
    rows = query.order_by(LoyaltyCard.card_number).limit(500).all()
    return success_response([_loyalty_card_payload(row) for row in rows])


@router.post("/loyalty-cards", status_code=status.HTTP_201_CREATED)
def create_loyalty_card(
    payload: LoyaltyCardCreate,
    current_user: User = Depends(require_loyalty_manage),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a loyalty card and optional opening balance transaction."""

    if session.query(LoyaltyCard).filter(LoyaltyCard.card_number == payload.card_number).one_or_none() is not None:
        raise HTTPException(status_code=409, detail=error_detail("DUPLICATE_CARD", "Loyalty card number already exists."))
    if payload.counterparty_id is not None:
        _get_or_404(session, Counterparty, payload.counterparty_id, "Counterparty")
    values = payload.model_dump()
    opening_balance = money(values.pop("balance_tmt"))
    card = LoyaltyCard(balance_tmt=Decimal("0.00"), **values)
    session.add(card)
    session.flush()
    if opening_balance > Decimal("0.00"):
        _post_loyalty_transaction(
            session,
            card,
            transaction_type="opening_balance",
            amount_tmt=opening_balance,
            doc_type="loyalty_card",
            doc_id=card.id,
            note="Opening balance",
            user_id=current_user.id,
        )
    session.commit()
    return success_response(_loyalty_card_payload(card))


@router.patch("/loyalty-cards/{card_id}")
def update_loyalty_card(
    card_id: int,
    payload: LoyaltyCardUpdate,
    _: User = Depends(require_loyalty_manage),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Patch loyalty card metadata."""

    card = _get_or_404(session, LoyaltyCard, card_id, "Loyalty card")
    updates = _updates(payload)
    if updates.get("counterparty_id") is not None:
        _get_or_404(session, Counterparty, int(updates["counterparty_id"]), "Counterparty")
    for key, value in updates.items():
        setattr(card, key, value)
    session.commit()
    return success_response(_loyalty_card_payload(card))


@router.get("/loyalty-cards/{card_id}/transactions")
def list_loyalty_transactions(
    card_id: int,
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List loyalty movements for one card."""

    _get_or_404(session, LoyaltyCard, card_id, "Loyalty card")
    rows = (
        session.query(LoyaltyTransaction)
        .options(selectinload(LoyaltyTransaction.loyalty_card))
        .filter(LoyaltyTransaction.loyalty_card_id == card_id)
        .order_by(LoyaltyTransaction.id.desc())
        .limit(500)
        .all()
    )
    return success_response([_loyalty_transaction_payload(row) for row in rows])


@router.post("/loyalty-cards/{card_id}/adjust", status_code=status.HTTP_201_CREATED)
def adjust_loyalty_card(
    card_id: int,
    payload: LoyaltyAdjustmentCreate,
    current_user: User = Depends(require_loyalty_adjust),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Post a manual loyalty-card balance adjustment."""

    card = _get_or_404(session, LoyaltyCard, card_id, "Loyalty card")
    transaction = _post_loyalty_transaction(
        session,
        card,
        transaction_type="manual_adjustment",
        amount_tmt=money(payload.amount_tmt),
        doc_type="loyalty_card",
        doc_id=card.id,
        note=payload.note or "Manual adjustment",
        user_id=current_user.id,
    )
    session.commit()
    return success_response(_loyalty_transaction_payload(transaction))


@router.get("/prices/current")
def get_current_price(
    price_list_id: int | None = Query(default=None),
    counterparty_id: int | None = Query(default=None),
    product_id: int | None = Query(default=None),
    service_id: int | None = Query(default=None),
    on_date: date | None = Query(default=None),
    _: User = Depends(require_pricing_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Return the current price for one product or service."""

    if (product_id is None) == (service_id is None):
        raise HTTPException(status_code=400, detail=error_detail("INVALID_TARGET", "Exactly one of product_id or service_id is required."))
    if price_list_id is None and counterparty_id is not None:
        counterparty = _get_or_404(session, Counterparty, counterparty_id, "Counterparty")
        price_list_id = counterparty.price_list_id
    if price_list_id is None:
        default_list = session.query(PriceList).filter(PriceList.is_default.is_(True), PriceList.is_active.is_(True)).one_or_none()
        if default_list is not None:
            price_list_id = default_list.id
    if price_list_id is None:
        raise HTTPException(status_code=404, detail=error_detail("PRICE_LIST_NOT_FOUND", "No price list selected and no default price list exists."))
    effective_date = on_date or date.today()
    query = (
        session.query(PriceListItem)
        .options(selectinload(PriceListItem.product), selectinload(PriceListItem.service), selectinload(PriceListItem.uom))
        .filter(
            PriceListItem.price_list_id == price_list_id,
            PriceListItem.valid_from <= effective_date,
            or_(PriceListItem.valid_to.is_(None), PriceListItem.valid_to >= effective_date),
        )
    )
    query = query.filter(PriceListItem.product_id == product_id) if product_id is not None else query.filter(PriceListItem.service_id == service_id)
    item = query.order_by(PriceListItem.valid_from.desc(), PriceListItem.id.desc()).first()
    if item is None:
        raise HTTPException(status_code=404, detail=error_detail("PRICE_NOT_FOUND", "No active price was found."))
    return success_response(_price_item_payload(item))


@router.get("/purchase-orders")
def list_purchase_orders(
    _: User = Depends(require_purchase_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List recent purchase orders."""

    rows = _order_query(session).order_by(PurchaseOrder.id.desc()).limit(200).all()
    return success_response([_order_payload(row) for row in rows])


@router.post("/purchase-orders", status_code=status.HTTP_201_CREATED)
def create_purchase_order(
    payload: PurchaseOrderCreate,
    current_user: User = Depends(require_purchase_order_create),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a draft supplier purchase order."""

    counterparty = _ensure_active_counterparty(session, payload.counterparty_id)
    _ensure_supplier(counterparty)
    _ensure_active_warehouse(session, payload.warehouse_id)
    _ensure_active_currency(session, payload.currency_id)
    doc_number = payload.doc_number or generate_doc_number(session, PurchaseOrder, "PO")
    if session.query(PurchaseOrder).filter(PurchaseOrder.doc_number == doc_number).one_or_none() is not None:
        raise HTTPException(status_code=409, detail=error_detail("DUPLICATE_DOC_NUMBER", "Purchase order number already exists."))
    order = PurchaseOrder(
        doc_number=doc_number,
        doc_date=payload.doc_date or date.today(),
        counterparty_id=payload.counterparty_id,
        warehouse_id=payload.warehouse_id,
        currency_id=payload.currency_id,
        currency_rate=payload.currency_rate,
        note=payload.note,
        created_by_user_id=current_user.id,
    )
    session.add(order)
    session.flush()
    for line_payload in payload.lines:
        order.lines.append(_create_order_line(session, order, line_payload))
    _recalculate_order_totals(order)
    session.commit()
    return success_response(_order_payload(_refresh_order(session, order.id)))


@router.get("/purchase-orders/{order_id}")
def get_purchase_order(
    order_id: int,
    _: User = Depends(require_purchase_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Return one purchase order."""

    return success_response(_order_payload(_refresh_order(session, order_id)))


@router.put("/purchase-orders/{order_id}")
def update_purchase_order(
    order_id: int,
    payload: PurchaseOrderUpdate,
    _: User = Depends(require_purchase_order_edit),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Update an unreceived purchase order."""

    order = _refresh_order(session, order_id)
    if order.status not in {"draft", "sent"}:
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Only draft or sent purchase orders can be edited."))
    if any(qty4(line.quantity_received) > Decimal("0.0000") for line in order.lines):
        raise HTTPException(status_code=400, detail=error_detail("ORDER_HAS_RECEIPTS", "Purchase order with received quantities cannot be edited."))
    updates = _updates(payload)
    if updates.get("counterparty_id") is not None:
        counterparty = _ensure_active_counterparty(session, int(updates["counterparty_id"]))
        _ensure_supplier(counterparty)
    if updates.get("warehouse_id") is not None:
        _ensure_active_warehouse(session, int(updates["warehouse_id"]))
    if updates.get("currency_id") is not None:
        _ensure_active_currency(session, int(updates["currency_id"]))
    for key in ("doc_date", "counterparty_id", "warehouse_id", "currency_id", "currency_rate", "note"):
        if key in updates:
            setattr(order, key, updates[key])
    if payload.lines is not None:
        order.lines.clear()
        session.flush()
        for line_payload in payload.lines:
            order.lines.append(_create_order_line(session, order, line_payload))
    _recalculate_order_totals(order)
    session.commit()
    return success_response(_order_payload(_refresh_order(session, order.id)))


@router.post("/purchase-orders/{order_id}/send")
def send_purchase_order(
    order_id: int,
    current_user: User = Depends(require_purchase_order_edit),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Mark a draft purchase order as sent to the supplier."""

    order = _refresh_order(session, order_id)
    if order.status != "draft":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Only draft purchase orders can be sent."))
    order.status = "sent"
    order.sent_by_user_id = current_user.id
    order.sent_at = now_utc()
    session.commit()
    return success_response(_order_payload(_refresh_order(session, order.id)))


@router.post("/purchase-orders/{order_id}/cancel")
def cancel_purchase_order(
    order_id: int,
    current_user: User = Depends(require_purchase_order_cancel),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Cancel a purchase order that has not received goods."""

    order = _refresh_order(session, order_id)
    if order.status == "cancelled":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Purchase order is already cancelled."))
    if any(qty4(line.quantity_received) > Decimal("0.0000") for line in order.lines):
        raise HTTPException(status_code=400, detail=error_detail("ORDER_HAS_RECEIPTS", "Cancel linked purchase invoices before cancelling this order."))
    order.status = "cancelled"
    order.cancelled_by_user_id = current_user.id
    order.cancelled_at = now_utc()
    session.commit()
    return success_response(_order_payload(_refresh_order(session, order.id)))


@router.get("/purchase-invoices")
def list_purchase_invoices(
    _: User = Depends(require_purchase_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List recent purchase invoices."""

    rows = _invoice_query(session).order_by(PurchaseInvoice.id.desc()).limit(200).all()
    return success_response([_invoice_payload(row) for row in rows])


@router.post("/purchase-invoices", status_code=status.HTTP_201_CREATED)
def create_purchase_invoice(
    payload: PurchaseInvoiceCreate,
    current_user: User = Depends(require_purchase_create),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a draft purchase invoice."""

    counterparty = _ensure_active_counterparty(session, payload.counterparty_id)
    _ensure_supplier(counterparty)
    _ensure_active_warehouse(session, payload.warehouse_id)
    _ensure_active_currency(session, payload.currency_id)
    if payload.purchase_order_id is not None:
        order = _refresh_order(session, payload.purchase_order_id)
        if order.status == "cancelled":
            raise HTTPException(status_code=400, detail=error_detail("ORDER_CANCELLED", "Purchase order is cancelled."))
        if order.counterparty_id != payload.counterparty_id or order.warehouse_id != payload.warehouse_id or order.currency_id != payload.currency_id:
            raise HTTPException(status_code=400, detail=error_detail("ORDER_HEADER_MISMATCH", "Invoice header differs from the linked purchase order."))
        if any(line.purchase_order_line_id is None for line in payload.lines):
            raise HTTPException(status_code=400, detail=error_detail("ORDER_LINE_REQUIRED", "Every invoice line linked to an order must reference a purchase order line."))
    if payload.return_invoice_id is not None:
        source_invoice = _get_or_404(session, PurchaseInvoice, payload.return_invoice_id, "Return source invoice")
        if source_invoice.status != "posted" or source_invoice.is_return:
            raise HTTPException(status_code=400, detail=error_detail("INVALID_RETURN_SOURCE", "Supplier returns must reference a posted purchase invoice."))
    doc_number = payload.doc_number or generate_doc_number(session, PurchaseInvoice, "PIN")
    if session.query(PurchaseInvoice).filter(PurchaseInvoice.doc_number == doc_number).one_or_none() is not None:
        raise HTTPException(status_code=409, detail=error_detail("DUPLICATE_DOC_NUMBER", "Purchase invoice number already exists."))
    invoice = PurchaseInvoice(
        doc_number=doc_number,
        doc_date=payload.doc_date or date.today(),
        purchase_order_id=payload.purchase_order_id,
        counterparty_id=payload.counterparty_id,
        warehouse_id=payload.warehouse_id,
        currency_id=payload.currency_id,
        currency_rate=payload.currency_rate,
        expiry_note=payload.expiry_note,
        is_return=payload.is_return,
        return_invoice_id=payload.return_invoice_id,
        note=payload.note,
        created_by_user_id=current_user.id,
    )
    session.add(invoice)
    session.flush()
    for line_payload in payload.lines:
        invoice.lines.append(_create_invoice_line(session, invoice, line_payload))
    invoice.total_amount_cur = money(sum((line.amount_cur for line in invoice.lines), Decimal("0")))
    invoice.total_amount_tmt = money(sum((line.amount_tmt for line in invoice.lines), Decimal("0")))
    if invoice.is_return:
        invoice.total_amount_cur = -invoice.total_amount_cur
        invoice.total_amount_tmt = -invoice.total_amount_tmt
    session.commit()
    return success_response(_invoice_payload(_refresh_invoice(session, invoice.id)))


@router.post("/purchase-invoices/return", status_code=status.HTTP_201_CREATED)
def create_purchase_return_invoice(
    payload: PurchaseInvoiceCreate,
    current_user: User = Depends(require_purchase_return),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create a supplier-return purchase invoice."""

    return create_purchase_invoice(payload.model_copy(update={"is_return": True}), current_user, session)


@router.get("/purchase-invoices/{invoice_id}")
def get_purchase_invoice(
    invoice_id: int,
    _: User = Depends(require_purchase_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Return one purchase invoice."""

    return success_response(_invoice_payload(_refresh_invoice(session, invoice_id)))


@router.post("/purchase-invoices/{invoice_id}/post")
def post_purchase_invoice(
    invoice_id: int,
    current_user: User = Depends(require_purchase_post),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Post a purchase invoice to stock and payable debt."""

    invoice = _refresh_invoice(session, invoice_id)
    if invoice.status != "draft":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Only draft purchase invoices can be posted."))
    try:
        for line in invoice.lines:
            if line.product_id is None:
                continue
            balance = get_or_create_balance(session, invoice.warehouse_id, line.product_id, line.uom_id)
            line.avg_cost_before = price(balance.avg_cost_tmt)
            movement = post_stock_movement(
                session,
                warehouse_id=invoice.warehouse_id,
                product_id=line.product_id,
                uom_id=line.uom_id,
                movement_type=("purchase_return" if invoice.is_return else "purchase"),
                document_type="purchase_invoice",
                document_id=invoice.id,
                quantity_delta=(-line.quantity if invoice.is_return else line.quantity),
                unit_cost_tmt=line.price_tmt,
                user_id=current_user.id,
            )
            balance_after = get_or_create_balance(session, invoice.warehouse_id, line.product_id, line.uom_id)
            line.avg_cost_after = price(balance_after.avg_cost_tmt if movement else line.avg_cost_before)
    except WarehouseBusinessError as exc:
        session.rollback()
        raise HTTPException(status_code=400, detail=error_detail(exc.code, str(exc), exc.details)) from exc

    invoice.status = "posted"
    invoice.posted_by_user_id = current_user.id
    invoice.posted_at = now_utc()
    post_debt_entry(
        session,
        counterparty_id=invoice.counterparty_id,
        debt_type="payable",
        doc_type="purchase_invoice",
        doc_id=invoice.id,
        doc_number=invoice.doc_number,
        doc_date=_doc_datetime(invoice.doc_date),
        amount_tmt=invoice.total_amount_tmt,
        currency_id=invoice.currency_id,
        amount_cur=invoice.total_amount_cur,
        note="Purchase invoice posted",
        user_id=current_user.id,
    )
    _apply_invoice_to_order(session, invoice, direction=1)
    session.commit()
    return success_response(_invoice_payload(_refresh_invoice(session, invoice.id)))


@router.post("/purchase-invoices/{invoice_id}/cancel")
def cancel_purchase_invoice(
    invoice_id: int,
    current_user: User = Depends(require_purchase_cancel),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Cancel a draft purchase invoice or reverse a posted one."""

    invoice = _refresh_invoice(session, invoice_id)
    if invoice.status == "cancelled":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Purchase invoice is already cancelled."))
    if invoice.status == "draft":
        invoice.status = "cancelled"
        session.commit()
        return success_response(_invoice_payload(_refresh_invoice(session, invoice.id)))
    if invoice.status != "posted":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Only draft or posted purchase invoices can be cancelled."))
    try:
        for line in invoice.lines:
            if line.product_id is None:
                continue
            post_stock_movement(
                session,
                warehouse_id=invoice.warehouse_id,
                product_id=line.product_id,
                uom_id=line.uom_id,
                movement_type=("purchase_return_cancel" if invoice.is_return else "purchase_cancel"),
                document_type="purchase_invoice",
                document_id=invoice.id,
                quantity_delta=(line.quantity if invoice.is_return else -line.quantity),
                unit_cost_tmt=line.price_tmt,
                user_id=current_user.id,
            )
    except WarehouseBusinessError as exc:
        session.rollback()
        raise HTTPException(status_code=400, detail=error_detail(exc.code, str(exc), exc.details)) from exc

    invoice.status = "cancelled"
    post_debt_entry(
        session,
        counterparty_id=invoice.counterparty_id,
        debt_type="payable",
        doc_type="purchase_invoice",
        doc_id=invoice.id,
        doc_number=invoice.doc_number,
        doc_date=now_utc(),
        amount_tmt=-invoice.total_amount_tmt,
        currency_id=invoice.currency_id,
        amount_cur=-invoice.total_amount_cur,
        note="Purchase invoice cancelled",
        user_id=current_user.id,
    )
    _apply_invoice_to_order(session, invoice, direction=-1)
    session.commit()
    return success_response(_invoice_payload(_refresh_invoice(session, invoice.id)))


@router.get("/debt-ledger")
def list_debt_ledger(
    counterparty_id: int | None = Query(default=None),
    debt_type: str | None = Query(default=None),
    _: User = Depends(require_debt_view),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """List recent debt ledger entries."""

    query = session.query(DebtLedger).options(selectinload(DebtLedger.counterparty))
    if counterparty_id is not None:
        query = query.filter(DebtLedger.counterparty_id == counterparty_id)
    if debt_type is not None:
        query = query.filter(DebtLedger.debt_type == debt_type)
    rows = query.order_by(DebtLedger.id.desc()).limit(500).all()
    return success_response([_debt_payload(row) for row in rows])


@router.post("/payments", status_code=status.HTTP_201_CREATED)
def create_payment(
    payload: PaymentCreate,
    current_user: User = Depends(require_payment_create),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Create and post a payment document."""

    _ensure_active_counterparty(session, payload.counterparty_id)
    if payload.currency_id is not None:
        _ensure_active_currency(session, payload.currency_id)
    if payload.cash_shift_id is not None:
        shift = _get_or_404(session, CashShift, payload.cash_shift_id, "Cash shift")
        if shift.status != "open":
            raise HTTPException(status_code=400, detail=error_detail("SHIFT_CLOSED", "Payment cash shift is not open."))
    allocation_total = money(sum((allocation.allocated_amount for allocation in payload.allocations), Decimal("0")))
    if allocation_total > money(payload.amount_tmt):
        raise HTTPException(status_code=400, detail=error_detail("ALLOCATION_EXCEEDS_PAYMENT", "Allocations cannot exceed payment amount."))
    doc_number = payload.doc_number or generate_doc_number(session, Payment, "PAY")
    if session.query(Payment).filter(Payment.doc_number == doc_number).one_or_none() is not None:
        raise HTTPException(status_code=409, detail=error_detail("DUPLICATE_DOC_NUMBER", "Payment number already exists."))
    payment = Payment(
        doc_number=doc_number,
        doc_date=payload.doc_date or now_utc(),
        counterparty_id=payload.counterparty_id,
        direction=payload.direction,
        payment_method=payload.payment_method,
        amount_tmt=money(payload.amount_tmt),
        currency_id=payload.currency_id,
        amount_cur=money(payload.amount_cur) if payload.amount_cur is not None else None,
        currency_rate=payload.currency_rate,
        cash_shift_id=payload.cash_shift_id,
        note=payload.note,
        created_by_user_id=current_user.id,
    )
    session.add(payment)
    session.flush()
    for allocation in payload.allocations:
        if allocation.doc_type == "purchase_invoice":
            invoice = _get_or_404(session, PurchaseInvoice, allocation.doc_id, "Purchase invoice")
            if invoice.counterparty_id != payment.counterparty_id:
                raise HTTPException(status_code=400, detail=error_detail("ALLOCATION_COUNTERPARTY_MISMATCH", "Allocation document belongs to another counterparty."))
        elif allocation.doc_type == "sale":
            sale = _get_or_404(session, Sale, allocation.doc_id, "Sale")
            if sale.counterparty_id != payment.counterparty_id:
                raise HTTPException(status_code=400, detail=error_detail("ALLOCATION_COUNTERPARTY_MISMATCH", "Allocation document belongs to another counterparty."))
        else:
            raise HTTPException(status_code=400, detail=error_detail("UNSUPPORTED_ALLOCATION", "Only purchase_invoice and sale allocations are supported."))
        payment.allocations.append(
            PaymentAllocation(
                doc_type=allocation.doc_type,
                doc_id=allocation.doc_id,
                allocated_amount=money(allocation.allocated_amount),
            )
        )
        if allocation.doc_type == "purchase_invoice":
            update_purchase_invoice_payment_status(session, invoice)

    debt_type = "payable" if payment.direction == "outgoing" else "receivable"
    post_debt_entry(
        session,
        counterparty_id=payment.counterparty_id,
        debt_type=debt_type,
        doc_type="payment",
        doc_id=payment.id,
        doc_number=payment.doc_number,
        doc_date=payment.doc_date,
        amount_tmt=-payment.amount_tmt,
        currency_id=payment.currency_id,
        amount_cur=payment.amount_cur,
        note="Payment posted",
        user_id=current_user.id,
    )
    session.commit()
    for allocation in payment.allocations:
        if allocation.doc_type == "purchase_invoice":
            invoice = session.get(PurchaseInvoice, allocation.doc_id)
            if invoice is not None:
                update_purchase_invoice_payment_status(session, invoice)
    session.commit()
    refreshed = (
        session.query(Payment)
        .options(selectinload(Payment.counterparty), selectinload(Payment.allocations))
        .filter(Payment.id == payment.id)
        .one()
    )
    return success_response(_payment_payload(refreshed))


@router.post("/payments/{payment_id}/cancel")
def cancel_payment(
    payment_id: int,
    current_user: User = Depends(require_payment_cancel),
    session: Session = Depends(get_db),
) -> dict[str, Any]:
    """Cancel a posted payment and reverse its debt effect."""

    payment = (
        session.query(Payment)
        .options(selectinload(Payment.counterparty), selectinload(Payment.allocations))
        .filter(Payment.id == payment_id)
        .one_or_none()
    )
    if payment is None:
        raise HTTPException(status_code=404, detail=error_detail("NOT_FOUND", "Payment not found."))
    if payment.status != "posted":
        raise HTTPException(status_code=400, detail=error_detail("INVALID_STATUS", "Only posted payments can be cancelled."))
    payment.status = "cancelled"
    payment.cancelled_by_user_id = current_user.id
    payment.cancelled_at = now_utc()
    debt_type = "payable" if payment.direction == "outgoing" else "receivable"
    post_debt_entry(
        session,
        counterparty_id=payment.counterparty_id,
        debt_type=debt_type,
        doc_type="payment",
        doc_id=payment.id,
        doc_number=payment.doc_number,
        doc_date=now_utc(),
        amount_tmt=payment.amount_tmt,
        currency_id=payment.currency_id,
        amount_cur=payment.amount_cur,
        note="Payment cancelled",
        user_id=current_user.id,
    )
    for allocation in payment.allocations:
        if allocation.doc_type == "purchase_invoice":
            invoice = session.get(PurchaseInvoice, allocation.doc_id)
            if invoice is not None:
                update_purchase_invoice_payment_status(session, invoice)
    session.commit()
    return success_response(_payment_payload(payment))
